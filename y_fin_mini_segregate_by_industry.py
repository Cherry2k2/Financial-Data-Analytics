import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Read the existing CSV file containing scraped company data
df = pd.read_csv("company_data.csv")

# Convert market cap values from string format to numerical format
def convert_market_cap(value):
    multiplier = 1
    if isinstance(value, str):
        if value.endswith('T'):
            multiplier = 1e12
        elif value.endswith('B'):
            multiplier = 1e9
        elif value.endswith('M'):
            multiplier = 1e6
        elif value.endswith('k'):
            multiplier = 1e3
        return float(value[:-1]) * multiplier
    return value

# Convert enterprise value values from string format to numerical format
def convert_enterprise_value(value):
    multiplier = 1
    if isinstance(value, str):
        if value.endswith('T'):
            multiplier = 1e12
        elif value.endswith('B'):
            multiplier = 1e9
        elif value.endswith('M'):
            multiplier = 1e6
        elif value.endswith('k'):
            multiplier = 1e3
        return float(value[:-1]) * multiplier
    return value

# Convert PB values from string format to numerical format
def convert_pb_value(value):
    if isinstance(value, str):
        return float(value)
    return value

# Convert number of employees values from string format to numerical format
def convert_employees(value):
    if isinstance(value, str):
        return int(value.replace(',', ''))
    return value

# Convert Trailing P/E values from string format to numerical format
def convert_pe_value(value):
    try:
        return float(value)
    except ValueError:
        return None

df['Market Cap'] = df['Market Cap'].apply(convert_market_cap)
df['Enterprise Value'] = df['Enterprise Value'].apply(convert_enterprise_value)
df['PB'] = df['PB'].apply(convert_pb_value)
df['No. of employees'] = df['No. of employees'].apply(convert_employees)
df['Trailing P/E'] = df['Trailing P/E'].apply(convert_pe_value)

# Function to check if a share price is close to its 52-week high
def is_close_to_52_week_high(row):
    return row['Indicator'] == 'Close to 52 week High'

# Group the data by 'Industry' and sort the groups
industry_groups = df.groupby('Industry')

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Create a Font object for bold text
bold_font = Font(bold=True)

# Define colors for highlighting
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue

# Start writing the data from the first row
row_idx = 1

# Iterate through each industry group
for industry, data in industry_groups:
    # Write the industry name to the Excel sheet
    ws.cell(row=row_idx, column=1, value=industry).font = bold_font
    row_idx += 1
    
    # Write the column headers for each industry
    col_idx = 1
    for header in df.columns:
        if header not in ['Sector', 'Industry']:  # Exclude 'Sector' and 'Industry'
            ws.cell(row=row_idx, column=col_idx, value=header).font = bold_font
            col_idx += 1
    
    row_idx += 1
    
    # Find the index of the row with the highest market cap value
    max_market_cap_index = data['Market Cap'].idxmax()
    
    # Find the index of the row with the highest enterprise value
    max_enterprise_value_index = data['Enterprise Value'].idxmax()
    
    # Find the index of the row with the highest PB value
    max_pb_index = data['PB'].idxmax()
    
    # Find the index of the row with the highest number of employees
    max_employees_index = data['No. of employees'].idxmax()
    
    # Find the index of the row with the lowest Trailing P/E value
    min_pe_index = data['Trailing P/E'].idxmin()
    
    # Find the index of the row with the lowest Beta value
    min_beta_index = data['Beta'].idxmin()
    
    # Write the data for each company in the industry
    for idx, row in data.iterrows():
        col_idx = 1
        for col in df.columns:  # Iterate over columns directly from df.columns
            if col not in ['Sector', 'Industry']:  # Exclude 'Sector' and 'Industry'
                cell = ws.cell(row=row_idx, column=col_idx, value=row[col])
                if row['Indicator'] == 'Close to 52 week High' and col == 'Indicator':  # Highlight 'Close to 52 week High' in orange
                    cell.fill = orange_fill
                elif row['Indicator'] == 'Close to 52 week low' and col == 'Indicator':  # Highlight 'Close to 52 week Low' in blue
                    cell.fill = blue_fill
                elif idx == max_market_cap_index and col == 'Market Cap':  # Highlight cell with highest market cap
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == max_enterprise_value_index and col == 'Enterprise Value':  # Highlight cell with highest enterprise value
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == max_pb_index and col == 'PB':  # Highlight cell with highest PB value
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == max_employees_index and col == 'No. of employees':  # Highlight cell with highest number of employees
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == min_pe_index and col == 'Trailing P/E':  # Highlight cell with lowest Trailing P/E value
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif idx == min_beta_index and col == 'Beta':  # Highlight cell with lowest Beta value
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                if col in ['Market Cap', 'Enterprise Value']:  # Set number format for Market Cap and Enterprise Value columns
                    cell.number_format = '#,##0.00'
                col_idx += 1
        row_idx += 1
    
    # Add an empty row to separate industries
    row_idx += 1
    
    # Calculate average P/E, P/B, and Beta for the industry
    avg_pe = data['Trailing P/E'].mean()
    avg_pb = data['PB'].mean()
    avg_beta = data['Beta'].mean()
    
    # Write the average P/E, P/B, and Beta one after the other at the end of the industry
    ws.cell(row=row_idx, column=1, value="Average P/E:").font = bold_font
    ws.cell(row=row_idx, column=2, value=avg_pe).font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Average P/B:").font = bold_font
    ws.cell(row=row_idx, column=2, value=avg_pb).font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Average Beta:").font = bold_font
    ws.cell(row=row_idx, column=2, value=avg_beta).font = bold_font
    
    row_idx += 2  # Skip two rows before next industry data

# Calculate share price-related information for all industries
# Find the industry with the highest share price
max_industry = df.groupby('Industry')['Market Cap'].mean().idxmax()
ws.cell(row=row_idx, column=1, value="Industry with Highest Share Price:").font = bold_font
ws.cell(row=row_idx, column=2, value=max_industry).font = bold_font

# Find the number of shares close to a 52-week high in the industry with the highest share price
close_to_52_week_high_count = df[df.apply(is_close_to_52_week_high, axis=1)].groupby('Industry').size()
if max_industry in close_to_52_week_high_count:
    close_to_52_week_high = close_to_52_week_high_count[max_industry]
else:
    close_to_52_week_high = 0

ws.cell(row=row_idx + 1, column=1, value="Shares Close to 52 Week High in Highest Share Price Industry:").font = bold_font
ws.cell(row=row_idx + 1, column=2, value=close_to_52_week_high).font = bold_font

# Adjust column widths for Market Cap and Enterprise Value columns
ws.column_dimensions['A'].width = 20  # Assuming 'Market Cap' is in column A
ws.column_dimensions['D'].width = 25  # Assuming 'Market Cap' is in column D
ws.column_dimensions['E'].width = 25  # Assuming 'Enterprise Value' is in column E

# Save the Excel workbook
wb.save("company_data_segregated_by_industry.xlsx")