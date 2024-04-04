import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

# Read the existing CSV file containing scraped company data
df = pd.read_csv("company_data.csv")

# Convert market cap values from string format to numerical format
def convert_market_cap(value):
    multiplier = 1
    if isinstance(value, str):
        if value.endswith('T'):
            multiplier = 1e12
        elif value.endswith('B'):
            multiplier = 1e9
        elif value.endswith('M'):
            multiplier = 1e6
        elif value.endswith('k'):
            multiplier = 1e3
        return float(value[:-1]) * multiplier
    return value

# Convert enterprise value values from string format to numerical format
def convert_enterprise_value(value):
    multiplier = 1
    if isinstance(value, str):
        if value.endswith('T'):
            multiplier = 1e12
        elif value.endswith('B'):
            multiplier = 1e9
        elif value.endswith('M'):
            multiplier = 1e6
        elif value.endswith('k'):
            multiplier = 1e3
        return float(value[:-1]) * multiplier
    return value

# Convert PB values from string format to numerical format
def convert_pb_value(value):
    if isinstance(value, str):
        return float(value)
    return value

# Convert number of employees values from string format to numerical format
def convert_employees(value):
    if isinstance(value, str):
        return int(value.replace(',', ''))
    return value

# Convert Trailing P/E values from string format to numerical format
def convert_pe_value(value):
    try:
        return float(value)
    except ValueError:
        return None

df['Market Cap'] = df['Market Cap'].apply(convert_market_cap)
df['Enterprise Value'] = df['Enterprise Value'].apply(convert_enterprise_value)
df['PB'] = df['PB'].apply(convert_pb_value)
df['No. of employees'] = df['No. of employees'].apply(convert_employees)
df['Trailing P/E'] = df['Trailing P/E'].apply(convert_pe_value)

# Function to check if a share price is close to its 52-week high
def is_close_to_52_week_high(row):
    return row['Indicator'] == 'Close to 52 week High'

# Group the data by 'Sector' and sort the groups
sector_groups = df.groupby('Sector')

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Create a Font object for bold text
bold_font = Font(bold=True)

# Define colors for highlighting
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Blue

# Write the headers to the Excel sheet with bold formatting
headers = df.columns
col_idx = 1
for header in headers:
    if header not in ['Sector', 'Industry']:  # Exclude 'Sector' and 'Industry'
        ws.cell(row=1, column=col_idx, value=header).font = bold_font
        col_idx += 1

# Start writing the data from the second row
row_idx = 2

# Iterate through each sector group
for sector, data in sector_groups:
    # Write the sector name to the Excel sheet
    ws.cell(row=row_idx, column=1, value=sector).font = bold_font
    row_idx += 1
    
    # Write the column headers for each sector
    col_idx = 1
    for header in headers:
        if header not in ['Sector', 'Industry']:  # Exclude 'Sector' and 'Industry'
            ws.cell(row=row_idx, column=col_idx, value=header).font = bold_font
            col_idx += 1
    
    row_idx += 1
    
    # Calculate the Employee to Market Cap ratio
    data['Employee to Market Cap Ratio'] = data['No. of employees'] / data['Market Cap']
    
    # Find the index of the row with the highest market cap value
    max_market_cap_index = data['Market Cap'].idxmax()
    
    # Find the index of the row with the highest enterprise value
    max_enterprise_value_index = data['Enterprise Value'].idxmax()
    
    # Find the index of the row with the highest PB value
    max_pb_index = data['PB'].idxmax()
    
    # Find the index of the row with the highest number of employees
    max_employees_index = data['No. of employees'].idxmax()
    
    # Find the index of the row with the lowest Trailing P/E value
    min_pe_index = data['Trailing P/E'].idxmin()
    
    # Find the index of the row with the lowest Beta value
    min_beta_index = data['Beta'].idxmin()
    
    # Write the data for each company in the sector
    for idx, row in data.iterrows():
        col_idx = 1
        for col in headers:  # Iterate over columns directly from df.columns
            if col not in ['Sector', 'Industry']:  # Exclude 'Sector' and 'Industry'
                cell = ws.cell(row=row_idx, column=col_idx, value=row[col])
                if row['Indicator'] == 'Close to 52 week High' and col == 'Indicator':  # Highlight 'Close to 52 week High' in orange
                    cell.fill = orange_fill
                elif row['Indicator'] == 'Close to 52 week low' and col == 'Indicator':  # Highlight 'Close to 52 week Low' in blue
                    cell.fill = blue_fill
                elif idx == max_market_cap_index and col == 'Market Cap':  # Highlight cell with highest market cap
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == max_enterprise_value_index and col == 'Enterprise Value':  # Highlight cell with highest enterprise value
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == max_pb_index and col == 'PB':  # Highlight cell with highest PB value
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == max_employees_index and col == 'No. of employees':  # Highlight cell with highest number of employees
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                elif idx == min_pe_index and col == 'Trailing P/E':  # Highlight cell with lowest Trailing P/E value
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                elif idx == min_beta_index and col == 'Beta':  # Highlight cell with lowest Beta value
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                if col in ['Market Cap', 'Enterprise Value']:  # Set number format for Market Cap and Enterprise Value columns
                    cell.number_format = '#,##0.00'
                col_idx += 1
        row_idx += 1
        
    # Add an empty row to separate industries
    row_idx += 1    
    
    # Calculate average P/E, P/B, and beta for the sector
    avg_pe = data['Trailing P/E'].mean()
    avg_pb = data['PB'].mean()
    avg_beta = data['Beta'].mean()

    # Write the average P/E, P/B, beta, and employee to market cap ratio at the end of the sector
    ws.cell(row=row_idx, column=1, value="Average P/E:").font = bold_font
    ws.cell(row=row_idx, column=2, value=avg_pe).font = bold_font
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Average P/B:").font = bold_font
    ws.cell(row=row_idx, column=2, value=avg_pb).font = bold_font
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Average Beta:").font = bold_font
    ws.cell(row=row_idx, column=2, value=avg_beta).font = bold_font

    row_idx += 2

    # Visualizations for each sector
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))

    # Market Cap Distribution
    sns.histplot(data=data, x='Market Cap', bins=10, ax=axes[0, 0])
    axes[0, 0].set_title('Market Cap Distribution')

    # Trailing P/E Distribution
    sns.histplot(data=data, x='Trailing P/E', bins=10, ax=axes[0, 1])
    axes[0, 1].set_title('Trailing P/E Distribution')

    # Share Price Trends
    for ticker, share_price in data.groupby('Ticker')['Share Price']:
        axes[1, 0].plot(share_price, label=ticker)
    axes[1, 0].set_title('Share Price Trends')
    axes[1, 0].legend()

    # # Correlation Heatmap
    # sns.heatmap(data.corr(), annot=True, cmap='coolwarm', fmt=".2f", ax=axes[1, 1])
    # axes[1, 1].set_title('Correlation Heatmap')

    # Adjust layout
    plt.tight_layout()

    # Convert the plot to an image and insert it into the Excel sheet
    img = Image(fig)
    ws.add_image(img, f"A{row_idx}")

    # Update row index for the next sector
    row_idx += 20

# Save the Excel workbook
wb.save("company_data_segregated_by_sector.xlsx")